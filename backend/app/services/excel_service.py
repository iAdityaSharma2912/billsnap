"""
Excel export service using openpyxl.
Bold headers, auto column widths, filters enabled, saved to
Exports/YYYY/Month/.
"""
import os
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.paths import APP_DATA_DIR, EXPORTS_DIR

# Kept as BASE_DIR for backward compatibility — see pdf_service.py for why.
BASE_DIR = APP_DATA_DIR

HEADERS = ["Invoice #", "Date", "Company", "Customer", "Amount", "Mode", "Items"]


def export_invoices_to_excel(invoices: List) -> str:
    """invoices: list of Invoice ORM objects (with .customer and .items loaded)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    ws.append(HEADERS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for inv in invoices:
        items_summary = ", ".join(f"{it.item_name} x{it.qty:g}" for it in inv.items)
        ws.append([
            inv.invoice_number,
            inv.invoice_date.strftime("%Y-%m-%d"),
            inv.company.value if hasattr(inv.company, "value") else inv.company,
            inv.customer.name if inv.customer else "",
            inv.grand_total,
            inv.payment_mode,
            items_summary,
        ])

    # auto column widths
    for col_idx, header in enumerate(HEADERS, start=1):
        max_len = len(header)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

    now = datetime.now()
    out_dir = os.path.join(EXPORTS_DIR, str(now.year), now.strftime("%B"))
    os.makedirs(out_dir, exist_ok=True)
    filename = f"Invoices-Export-{now.strftime('%Y%m%d-%H%M%S')}.xlsx"
    full_path = os.path.join(out_dir, filename)
    wb.save(full_path)

    return os.path.relpath(full_path, BASE_DIR)
