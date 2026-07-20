"""Tests for /export/* — generate the .xlsx, then download it as real bytes."""


def _make_customer(client, name="Test Customer", phone="9000000001"):
    r = client.post("/customer/create", json={"name": name, "phone": phone})
    return r.json()["id"]


def _make_invoice(client, company="SA"):
    customer_id = _make_customer(client)
    r = client.post("/invoice/create", json={
        "company": company, "customer_id": customer_id, "payment_mode": "Cash",
        "items": [{"item_name": "Export Test Item", "unit": "Pcs", "qty": 1, "rate": 100}],
    })
    return r.json()


def test_export_excel_generates_path(client):
    _make_invoice(client)
    r = client.post("/export/excel")
    assert r.status_code == 200
    assert r.json()["path"].endswith(".xlsx")
    assert r.json()["path"].startswith("exports")


def test_export_excel_download_returns_real_xlsx_bytes(client):
    """The actual bug being fixed: previously there was no way to get the
    bytes of the generated file at all — only a server-side path string."""
    _make_invoice(client)
    r = client.post("/export/excel")
    path = r.json()["path"]

    r2 = client.get("/export/excel/download", params={"path": path})
    assert r2.status_code == 200
    assert r2.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert len(r2.content) > 1000  # a real xlsx, not an empty/error response
    # Without this header, the browser displays/navigates instead of saving.
    assert "attachment" in r2.headers["content-disposition"]
    assert path.split("/")[-1] in r2.headers["content-disposition"]


def test_export_excel_download_missing_file_404(client):
    r = client.get("/export/excel/download", params={"path": "exports/2026/June/does-not-exist.xlsx"})
    assert r.status_code == 404


def test_export_excel_download_rejects_path_traversal(client):
    """path must stay inside exports/ — reject attempts to read arbitrary
    files elsewhere on disk (e.g. the database)."""
    r = client.get("/export/excel/download", params={"path": "../database/smartbill.db"})
    assert r.status_code == 400

    r2 = client.get("/export/excel/download", params={"path": "database/smartbill.db"})
    assert r2.status_code == 400


def test_export_excel_filters_by_company(client):
    _make_invoice(client, company="SA")
    _make_invoice(client, company="SE")

    r = client.post("/export/excel", params={"company": "SA"})
    assert r.status_code == 200
    path = r.json()["path"]

    # downloadable and contains only the filtered company's rows — verified
    # indirectly via openpyxl since the API only returns a path/bytes.
    r2 = client.get("/export/excel/download", params={"path": path})
    assert r2.status_code == 200

    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r2.content))
    ws = wb.active
    companies_in_sheet = {row[2] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    assert companies_in_sheet == {"SA"}
